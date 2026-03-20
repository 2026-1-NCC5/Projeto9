from PIL import Image
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

img = Image.open("logo-LE.png")
img = img.convert("RGB") 

matriz = np.array(img)

#print(matriz.shape)

altura, largura, _ = matriz.shape

# Converte em CSV
dados = []

for y in range(altura):
    for x in range(largura):
        r, g, b = matriz[y][x]
        dados.append([y, x, r, g, b])

df = pd.DataFrame(dados, columns=["y", "x", "R", "G", "B"])
df.to_csv("imagem.csv", index=False)

reconstruida = np.array(matriz, dtype=np.uint8)

img_reconstruida = Image.fromarray(reconstruida)
img_reconstruida.save("reconstruida.png")

#Converte em xlsx - Igual o Rod mostrou a primeira aula
img = Image.open("logo-LE.png").convert("RGB")
img = img.resize((100, 100))
matriz = np.array(img)

img = img.resize((100, 100))
matriz = np.array(img)
altura, largura, _ = matriz.shape

wb = Workbook()
ws = wb.active

# 1. AJUSTE DE DIMENSÕES
# Para o Excel, usamos 1 até altura (inclusive)
for i in range(1, altura + 1):
    ws.row_dimensions[i].height = 5

for j in range(1, largura + 1):
    col_letter = get_column_letter(j)
    ws.column_dimensions[col_letter].width = 1

# 2. PREENCHIMENTO DAS CORES
# Usamos range(altura), que vai de 0 a 99 (perfeito para a matriz)
for y in range(altura):
    for x in range(largura):
        # Acessa a matriz (0 a 99)
        r, g, b = matriz[y, x] 

        cor_hex = f"{r:02X}{g:02X}{b:02X}"

        fill = PatternFill(start_color=cor_hex,
                           end_color=cor_hex,
                           fill_type="solid")

        # No Excel, as células começam em 1, por isso somamos +1 aqui
        ws.cell(row=y + 1, column=x + 1).fill = fill

wb.save("imagem_excel.xlsx")

#Cria uma janela para comparar as imagens
plt.subplot(1, 2, 1)
plt.title("Original")
plt.imshow(img)
plt.axis("off")

plt.subplot(1, 2, 2)
plt.title("Reconstruída")
plt.imshow(img_reconstruida)
plt.axis("off")

plt.show()

